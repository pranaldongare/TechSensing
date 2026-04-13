import re
from typing import Any, Dict, List, Optional, Tuple, Union

import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd


def detect_merged_header_rows(
    file_path: str, sheet_name: str, header_row_idx: int, max_scan_rows: int = 10
) -> Union[List[int], int]:
    """
    Detect if the sheet has multi-level headers caused by merged cells.

    Checks for horizontally merged cells in the header region. If a merged cell
    spans multiple columns in the row at or near header_row_idx, the row below
    it likely contains sub-headers (e.g., "Budget" spans "Plan" and "Actual").

    Returns:
        List[int] if multi-level headers detected (e.g., [2, 3])
        int (header_row_idx) if single-level headers
    """
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return header_row_idx
        ws = wb[sheet_name]

        # Collect merged cell ranges in the header region
        merged_in_header = []
        for merged_range in ws.merged_cells.ranges:
            # merged_range has min_row, max_row, min_col, max_col (1-based)
            # header_row_idx is 0-based, so the header row in openpyxl is header_row_idx + 1
            header_row_1based = header_row_idx + 1

            # Check if this merge is in the header region (within a few rows of detected header)
            if (
                merged_range.min_row <= header_row_1based + 1
                and merged_range.min_row >= max(1, header_row_1based - 1)
            ):
                # Check if it spans multiple columns (horizontal merge = multi-level header)
                col_span = merged_range.max_col - merged_range.min_col + 1
                if col_span > 1:
                    merged_in_header.append(merged_range)

        wb.close()

        if not merged_in_header:
            return header_row_idx

        # We found horizontal merges in the header area.
        # Determine the parent and child header rows:
        # The merged cells are in the parent row, the row below has sub-headers.
        parent_rows = set()
        for mr in merged_in_header:
            parent_rows.add(mr.min_row)

        # Convert to 0-based for pandas
        parent_row_0based = min(parent_rows) - 1
        child_row_0based = parent_row_0based + 1

        # Validate: child row should exist and be within reasonable range
        if child_row_0based < max_scan_rows:
            return [parent_row_0based, child_row_0based]

        return header_row_idx

    except Exception as e:
        print(f"Error detecting merged headers for sheet {sheet_name}: {e}")
        return header_row_idx


def flatten_multiindex_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Flatten MultiIndex columns into single-level names.

    E.g., ('Budget', 'Plan') -> 'Budget_Plan'
          ('Budget', 'Actual') -> 'Budget_Actual'
          ('', 'Name') -> 'Name'
          ('Total', '') -> 'Total'
    """
    if not isinstance(df.columns, pd.MultiIndex):
        return df

    new_columns = []
    for col_tuple in df.columns:
        # Filter out empty/NaN/Unnamed parts
        parts = []
        for part in col_tuple:
            part_str = str(part).strip()
            if (
                part_str
                and part_str.lower() != "nan"
                and not part_str.startswith("Unnamed")
            ):
                parts.append(part_str)

        if parts:
            new_columns.append("_".join(parts))
        else:
            new_columns.append(f"Column_{len(new_columns)}")

    df.columns = new_columns
    return df


def deduplicate_columns(columns: List[str]) -> List[str]:
    """
    Append _2, _3, etc. to duplicate column names to ensure uniqueness.

    E.g., ['amount', 'name', 'amount'] -> ['amount', 'name', 'amount_2']
    """
    seen = {}
    result = []
    for col in columns:
        col_lower = col.lower()
        if col_lower in seen:
            seen[col_lower] += 1
            result.append(f"{col}_{seen[col_lower]}")
        else:
            seen[col_lower] = 1
            result.append(col)
    return result


def find_header_row(
    file_path: str, sheet_name: str, max_scan_rows: int = 20
) -> Tuple[int, Optional[str]]:
    """
    Heuristic to find the header row index and extraction of pre-header context.

    Returns:
        (header_index, context_text)
        header_index: 0-based index of the header row (to pass to pd.read_excel header=N)
        context_text: Text found above the header row (e.g. titles, dates)
    """
    try:
        # Read first N rows without header to inspect content
        df_preview = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            header=None,
            nrows=max_scan_rows,
            engine="openpyxl",
        )
    except Exception as e:
        print(f"Error reading preview for sheet {sheet_name}: {e}")
        return 0, None

    best_score = -1
    best_row_idx = 0
    row_scores = []

    # Heuristic scoring
    for idx, row in df_preview.iterrows():
        # Convert row to list of strings (handling NaNs)
        row_values = [
            str(val).strip() for val in row if pd.notna(val) and str(val).strip() != ""
        ]

        if not row_values:
            continue

        non_empty_count = len(row_values)
        unique_count = len(set(row_values))

        # Rule 1: Headers are rarely numeric
        string_ratio = (
            sum(1 for v in row_values if not v.replace(".", "", 1).isdigit())
            / non_empty_count
            if non_empty_count
            else 0
        )

        # Rule 2: Headers usually have high uniqueness (no repeated column names)
        uniqueness_score = unique_count / non_empty_count if non_empty_count else 0

        # Rule 3: Headers usually have more filled columns than metadata rows (like "Date: ...")
        fullness_score = non_empty_count / len(df_preview.columns)

        # Composite score
        # We value stringiness and fullness highly. Uniqueness is a strong signal too.
        score = (string_ratio * 0.4) + (uniqueness_score * 0.3) + (fullness_score * 0.3)

        # Penalty for very short rows (likely just a title "Sales Report")
        if non_empty_count < 2:
            score *= 0.1

        row_scores.append((idx, score))

        if score > best_score:
            best_score = score
            best_row_idx = idx

    # If no decent header found, default to 0
    if best_score < 0.3:
        best_row_idx = 0

    # Extract context (rows before the header)
    context_lines = []
    if best_row_idx > 0:
        # Re-slice checks
        pre_header_df = df_preview.iloc[:best_row_idx]
        for _, row in pre_header_df.iterrows():
            # Join non-null values with space
            text_row = " ".join(
                [
                    str(val).strip()
                    for val in row
                    if pd.notna(val) and str(val).strip() != ""
                ]
            )
            if text_row:
                context_lines.append(text_row)

    context_text = "\n".join(context_lines) if context_lines else None

    return best_row_idx, context_text


def extract_metadata_from_cell(cell) -> List[str]:
    """Helper to extract comments and semantic colors from an openpyxl cell."""
    metadata = []

    # Extract Comment
    if cell.comment:
        metadata.append(f"[Note: {cell.comment.text.strip()}]")

    # Extract Color (Simple heuristic for generic Red/Green/Yellow)
    # openpyxl colors are RGB (ARGB) often.
    if cell.fill and cell.fill.start_color and cell.fill.start_color.index:
        # Note: Handling theme colors is complex. We focus on explicit RGB for now.
        # This is basic; a full implementation would map theme indices.
        color = cell.fill.start_color.rgb
        # Ignore White/None/Transparent
        if color and color not in ["00000000", "FFFFFFFF", None]:
            # Rough mapping (optional)
            if str(color).startswith("FF00FF00"):
                metadata.append("[Status: Green]")
            elif str(color).startswith("FFFF0000"):
                metadata.append("[Status: Red]")
            # For now, just generic fallback
            else:
                # Reduce noise: don't tag every gray header
                pass

    return metadata


def enrich_dataframe_with_metadata(
    df: pd.DataFrame, file_path: str, sheet_name: str, header_row_idx: int
) -> pd.DataFrame:
    """
    Reloads the sheet using openpyxl directly to attach comments/colors to the values.
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            return df
        ws = wb[sheet_name]

        # Create a map of coordinate -> metadata
        meta_map = {}

        # openpyxl uses 1-based indexing.
        # header_row_idx is 0-based from pandas read without header?
        # If we passed header=N to read_excel, the data starts at N+2 (1-based header is N+1, data is N+2)
        # It's safer to iterate the used range of the worksheet matching the DataFrame shape.

        # Offset: header_row_idx is the 0-based index of the header.
        # So in Excel: Row = header_row_idx + 1 is the header. Data starts at header_row_idx + 2.
        start_row = header_row_idx + 2

        # We only care about data cells matching the DF columns
        # This can be slow for massive sheets, so maybe limit to top 1000 rows or prioritize comments?
        for i, row in enumerate(
            ws.iter_rows(min_row=start_row, values_only=False), start=0
        ):
            if i >= len(df):
                break  # Don't go past DF bounds

            for j, cell in enumerate(row):
                if j >= len(df.columns):
                    break

                meta = extract_metadata_from_cell(cell)
                if meta:
                    col_name = df.columns[j]
                    # Append metadata to the DataFrame value
                    original_val = df.iat[i, j]
                    # Avoid appending to NaNs if we want to keep them empty, or convert to string
                    if pd.notna(original_val):
                        df.iat[i, j] = f"{original_val} {' '.join(meta)}"
                    elif meta:
                        df.iat[i, j] = f"{' '.join(meta)}"

        return df

    except Exception as e:
        print(f"Metadata enrichment failed: {e}")
        return df
